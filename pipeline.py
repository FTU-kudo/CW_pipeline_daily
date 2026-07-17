"""
CW Pipeline - Incremental Mode (Full)
- Scrape Vietstock, OHLCV (vnstock), filter, export Excel & JSON
- Tích hợp lấy đường cong lợi suất TPCP từ HNX (cache CSV)
- Tự tính Black‑Scholes: sigma 252 phiên, lãi suất 10Y mới nhất
- Xuất dữ liệu JSON có giá lý thuyết Call & Put, Delta, Gamma, Vega, Theta
"""

import os, re, time, threading, requests, pandas as pd, numpy as np, json
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date, timedelta, timezone
from math import log, sqrt, exp, pi, erf
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────── HÀM TIỆN ÍCH TOÁN HỌC CHO BLACK‑SCHOLES ──────────────────────────
def norm_cdf(x: float) -> float:
    """Hàm phân phối tích luỹ chuẩn (dùng math.erf)."""
    return 0.5 * (1.0 + erf(x / np.sqrt(2.0)))

def norm_pdf(x: float) -> float:
    """Hàm mật độ xác suất chuẩn."""
    return np.exp(-0.5 * x**2) / np.sqrt(2.0 * pi)

def black_scholes_option(S: float, K: float, T: float, r: float, sigma: float,
                         ratio: float = 1.0, option_type: str = 'call') -> dict:
    """
    Tính giá quyền chọn kiểu Âu (call hoặc put) và các chỉ số phái sinh.
    S, K : giá cơ sở và giá thực hiện (VND)
    T    : thời gian đến đáo hạn (năm)
    r    : lãi suất phi rủi ro (dạng thập phân, vd 0.05 = 5%)
    sigma: độ biến động hàng năm
    ratio: tỷ lệ chuyển đổi (số CW cần để đổi 1 cổ phiếu)
    option_type: 'call' hoặc 'put'
    Trả về dict với bs_price (nghìn đồng), delta, gamma, vega, theta.
    """
    if S <= 0 or K <= 0 or T <= 0 or sigma <= 0:
        return {"bs_price": 0, "delta_bs": 0, "gamma": 0, "vega": 0, "theta": 0}

    d1 = (log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * sqrt(T))
    d2 = d1 - sigma * sqrt(T)

    if option_type.lower() == 'call':
        option_value = S * norm_cdf(d1) - K * exp(-r * T) * norm_cdf(d2)
        delta = norm_cdf(d1)
        theta = (- (S * norm_pdf(d1) * sigma) / (2 * sqrt(T))
                 - r * K * exp(-r * T) * norm_cdf(d2)) / 365.0
    else:  # put
        option_value = K * exp(-r * T) * norm_cdf(-d2) - S * norm_cdf(-d1)
        delta = norm_cdf(d1) - 1
        theta = (- (S * norm_pdf(d1) * sigma) / (2 * sqrt(T))
                 + r * K * exp(-r * T) * norm_cdf(-d2)) / 365.0

    bs_price = option_value / 1000.0 / ratio   # giá CW (nghìn đồng)
    gamma = norm_pdf(d1) / (S * sigma * sqrt(T)) / ratio
    vega  = S * norm_pdf(d1) * sqrt(T) / 100.0 / ratio

    return {
        "bs_price": round(bs_price, 3),
        "delta_bs": round(delta, 4),
        "gamma":    round(gamma, 6),
        "vega":     round(vega, 4),
        "theta":    round(theta, 6)
    }

def historical_volatility(prices: np.ndarray, window: int = 252) -> float:
    """
    Ước lượng độ biến động lịch sử (annualized) từ chuỗi giá đóng cửa.
    Sử dụng tối đa `window` phiên gần nhất.
    """
    prices = np.asarray(prices, dtype=float)
    prices = prices[prices > 0]
    if len(prices) < 2:
        return 0.0
    if len(prices) > window:
        prices = prices[-window:]
    log_returns = np.diff(np.log(prices))
    if len(log_returns) == 0:
        return 0.0
    return np.std(log_returns, ddof=1) * np.sqrt(252)

# ────────────────────────── THU THẬP LỢI SUẤT TPCP TỪ HNX ──────────────────────────
try:
    from curl_cffi import requests as cffi_requests
except ImportError:
    cffi_requests = None

try:
    from vnstock import Vnstock
except ImportError:
    Vnstock = None

import contextlib, io

def fetch_vnstock_yields(start_date: str, end_date: str) -> pd.DataFrame:
    """Lớp 1: thử dùng thư viện vnstock."""
    if Vnstock is None:
        return pd.DataFrame()
    try:
        api_key = os.getenv('VNSTOCK_API_KEY')
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            vs = Vnstock(token=api_key) if api_key else Vnstock()
            if hasattr(vs, 'bond'):
                bond_data = vs.bond(source='VCI')
                if hasattr(bond_data, 'bond_yield'):
                    df = bond_data.bond_yield(start_date=start_date, end_date=end_date)
                    if not df.empty and 'date' in df.columns:
                        df['date'] = pd.to_datetime(df['date'])
                        df = df.set_index('date').sort_index()
                        standard_tenors = ['1Y', '2Y', '3Y', '5Y', '7Y', '10Y', '15Y']
                        avail = [c for c in standard_tenors if c in df.columns]
                        if avail:
                            return df[avail].dropna()
    except Exception:
        pass
    return pd.DataFrame()

def fetch_public_api_yields(start_date: str, end_date: str) -> pd.DataFrame:
    """Lớp 2: placeholder cho API công khai."""
    return pd.DataFrame()

def scrape_trading_economics_10y() -> float:
    """Fallback lợi suất 10Y từ Trading Economics."""
    if cffi_requests is None:
        return 4.53
    try:
        url = "https://tradingeconomics.com/vietnam/government-bond-yield"
        r = cffi_requests.get(url, impersonate="chrome120", timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "html.parser")
            for tr in soup.find_all("tr"):
                txt = tr.get_text(" | ", strip=True)
                if "Vietnam 10Y" in txt:
                    parts = txt.split(" | ")
                    for p in parts[1:]:
                        try:
                            val = float(p.replace('%', '').strip())
                            if 1.0 <= val <= 10.0:
                                print(f"[+] Trading Economics 10Y: {val}%")
                                return val
                        except ValueError:
                            continue
    except Exception as e:
        print(f"[-] Trading Economics scrape error: {e}")
    return 4.53

def fetch_single_hnx_date(dt_str: str) -> dict | None:
    """Lấy đường cong lợi suất 1 ngày từ HNX."""
    if cffi_requests is None:
        return None
    try:
        dt = pd.to_datetime(dt_str)
        url = "https://www.hnx.vn/ModuleReportBonds/Bond_YieldCurve/SearchAndNextPageYieldCurveData"
        for offset in range(6):
            check_dt = dt - timedelta(days=offset)
            p_date = check_dt.strftime("%d/%m/%Y")
            try:
                r = cffi_requests.post(url, data={"pDate": p_date}, impersonate="chrome120",
                                       verify=False, timeout=12)
                soup = BeautifulSoup(r.text, "html.parser")
                yields = {}
                for tr in soup.find_all("tr"):
                    row = [td.get_text(strip=True) for td in tr.find_all(["th", "td"])]
                    if len(row) >= 4:
                        tenor = row[0].strip().lower()
                        val_str = row[3] if row[3] else row[1]
                        val_str = val_str.replace(",", ".").strip()
                        try:
                            val = float(val_str)
                            if tenor == "3 tháng": yields["3M"] = round(val, 3)
                            elif tenor == "6 tháng": yields["6M"] = round(val, 3)
                            elif tenor == "9 tháng": yields["9M"] = round(val, 3)
                            elif tenor == "1 năm": yields["1Y"] = round(val, 3)
                            elif tenor == "2 năm": yields["2Y"] = round(val, 3)
                            elif tenor == "3 năm": yields["3Y"] = round(val, 3)
                            elif tenor == "5 năm": yields["5Y"] = round(val, 3)
                            elif tenor == "7 năm": yields["7Y"] = round(val, 3)
                            elif tenor == "10 năm": yields["10Y"] = round(val, 3)
                            elif tenor == "15 năm": yields["15Y"] = round(val, 3)
                            elif tenor == "20 năm": yields["20Y"] = round(val, 3)
                        except ValueError:
                            pass
                if "10Y" in yields and "1Y" in yields:
                    yields["date"] = dt_str
                    return yields
            except Exception:
                pass
    except Exception:
        pass
    return None

def fetch_hnx_official_yields(start_date: str, end_date: str) -> pd.DataFrame:
    """Lớp 3: dữ liệu thực tế từ HNX, có cache CSV."""
    if cffi_requests is None:
        print("[!] curl_cffi chưa được cài đặt, không thể scrape HNX.")
        return pd.DataFrame()
    start_dt = pd.to_datetime(start_date)
    end_dt = pd.to_datetime(end_date)
    target_dates = pd.date_range(start=start_dt, end=end_dt, freq='B')
    if len(target_dates) == 0:
        target_dates = pd.date_range(start='2016-01-01', end=datetime.now(), freq='B')

    cache_path = os.path.join(CACHE_DIR, 'hnx_bond_yields.csv')
    df_cache = pd.DataFrame()
    if os.path.exists(cache_path):
        try:
            df_cache = pd.read_csv(cache_path)
            df_cache['date'] = pd.to_datetime(df_cache['date'])
            df_cache = df_cache.set_index('date').sort_index()
        except Exception as e:
            print(f"[-] Lỗi đọc cache HNX: {e}")

    existing_dates = set(df_cache.index.strftime('%Y-%m-%d')) if not df_cache.empty else set()
    missing_dates = [d.strftime('%Y-%m-%d') for d in target_dates if d.strftime('%Y-%m-%d') not in existing_dates]

    if missing_dates:
        print(f"[*] Đang tải {len(missing_dates)} ngày lợi suất mới từ HNX...")
        with ThreadPoolExecutor(max_workers=16) as executor:
            new_records = list(executor.map(fetch_single_hnx_date, missing_dates))
        new_records = [r for r in new_records if r is not None]
        if new_records:
            df_new = pd.DataFrame(new_records)
            df_new['date'] = pd.to_datetime(df_new['date'])
            df_new = df_new.set_index('date')
            df_cache = pd.concat([df_cache, df_new]).sort_index()
            df_cache = df_cache[~df_cache.index.duplicated(keep='last')]
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            df_cache.to_csv(cache_path, encoding='utf-8-sig')

    if df_cache.empty:
        return pd.DataFrame()
    df = df_cache.loc[start_dt:end_dt].copy()
    for t in ['3M', '6M', '9M', '1Y', '2Y', '3Y', '5Y', '7Y', '10Y', '15Y', '20Y']:
        if t not in df.columns:
            df[t] = np.nan
    df = df.ffill().bfill()
    return df

def fetch_bond_yields(start_date: str, end_date: str) -> pd.DataFrame:
    """Tổng hợp lợi suất TPCP Việt Nam, ưu tiên HNX."""
    print(f"[*] Thu thập dữ liệu lợi suất TPCP ({start_date} -> {end_date})...")
    df = fetch_vnstock_yields(start_date, end_date)
    if not df.empty and len(df) > 50:
        print(f"[+] Lớp 1 (vnstock): {len(df)} dòng")
        return df
    df = fetch_public_api_yields(start_date, end_date)
    if not df.empty and len(df) > 50:
        print(f"[+] Lớp 2 (public API): {len(df)} dòng")
        return df
    print("[+] Kích hoạt Lớp 3: HNX chính thức...")
    df = fetch_hnx_official_yields(start_date, end_date)
    print(f"[+] Hoàn tất lợi suất HNX: {len(df)} ngày")
    return df

# ══════════════════════════════════════════════════════════════════
# CẤU HÌNH (GIỮ NGUYÊN TỪ PIPELINE GỐC)
# ══════════════════════════════════════════════════════════════════
BASE_STOCKS = [
    "CACB","CDGC","CFPT","CHDB","CHPG","CLPB","CMBB","CMSN","CMWG","CSHB",
    "CSSB","CSTB","CTCB","CTPB","CVHM","CVIB","CVIC","CVJC","CVNM","CVPB",
    "CVRE","CPOW","CBVH","CBCM","CBSR",
]
YEARS        = ["23","24","25","26"]
MAX_ISSUANCE = 50
MAX_WORKERS  = 10
TIMEOUT      = 12
DELAY        = 0.1

OHLCV_START_DATE = "2023-01-01"
FILTER_DATE      = date(2024, 1, 2)
MAX_RETRIES      = 3        # FIX: giảm từ 5 → 3; worst-case 1 CW = 3×30s+2×4s = 98s thay vì 170s
RETRY_DELAY      = 4.0      # FIX: giảm từ 5.0 → 4.0s
REQUEST_DELAY    = 0.6

CACHE_DIR       = "output/cache"
VIETSTOCK_CACHE = f"{CACHE_DIR}/vietstock.parquet"
OHLCV_CACHE     = f"{CACHE_DIR}/ohlcv.parquet"
OUTPUT_FILE     = "output/cw_master.xlsx"

# ══════════════════════════════════════════════════════════════════
# CACHE HELPERS
# ══════════════════════════════════════════════════════════════════
def load_cache(path):
    if os.path.exists(path):
        df = pd.read_parquet(path)
        print(f"   Doc cache: {path}  ({len(df):,} dong)")
        return df
    print(f"   Chua co cache: {path}")
    return None

def save_cache(df, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if "time" in df.columns:
        df = df.copy()
        df["time"] = df["time"].astype(str)
    df.to_parquet(path, index=False)
    print(f"   Luu cache: {path}  ({len(df):,} dong)")

# ══════════════════════════════════════════════════════════════════
# VIETSTOCK SCRAPER (giữ nguyên)
# ══════════════════════════════════════════════════════════════════
BASE_URL = "https://finance.vietstock.vn"
HEADERS  = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "vi-VN,vi;q=0.9",
    "Referer": f"{BASE_URL}/chung-khoan-phai-sinh/chung-quyen.htm",
}

_ALL_TABLE_LABELS = sorted([
    "To chuc phat hanh CKCS","To chuc phat hanh CW",
    "Ngay giao dich cuoi cung","Ngay giao dich dau tien",
    "Khoi luong Niem yet","Khoi luong luu hanh",
    "Loai chung quyen","Kieu thuc hien",
    "TLCD dieu chinh","Gia TH dieu chinh","Ty le chuyen doi",
    "Ngay phat hanh","Ngay niem yet","Ngay dao han",
    "Gia phat hanh","Gia thuc hien","CK co so","Thoi han","Tai lieu",
    "To chuc phat hanh CKCS","To chuc phat hanh CW",
    "Phuong thuc thuc hien quyen","Ngay giao dich cuoi cung",
    "Ngay giao dich dau tien","Khoi luong Niem yet",
    "Khoi luong luu hanh","Loai chung quyen","Kieu thuc hien",
    "TLCD dieu chinh","Gia TH dieu chinh","Ty le chuyen doi",
    "Ngay phat hanh","Ngay niem yet","Ngay dao han",
    "Gia phat hanh","Gia thuc hien","CK co so","Thoi han","Tai lieu",
    # UTF-8
    "Tổ chức phát hành CKCS","Tổ chức phát hành CW",
    "Phương thức thực hiện quyền","Ngày giao dịch cuối cùng",
    "Ngày giao dịch đầu tiên","Khối lượng Niêm yết",
    "Khối lượng lưu hành","Loại chứng quyền","Kiểu thực hiện",
    "TLCĐ điều chỉnh","Giá TH điều chỉnh","Tỷ lệ chuyển đổi",
    "Ngày phát hành","Ngày niêm yết","Ngày đáo hạn",
    "Giá phát hành","Giá thực hiện","CK cơ sở","Thời hạn","Tài liệu",
], key=len, reverse=True)

_LABEL_SPLIT_RE = re.compile(
    r'('+  '|'.join(re.escape(l) for l in _ALL_TABLE_LABELS) + r')\s*:',
    re.UNICODE
)
_thread_local = threading.local()

def get_session():
    if not hasattr(_thread_local,"session"):
        s = requests.Session(); s.headers.update(HEADERS)
        _thread_local.session = s
    return _thread_local.session

def clean_cell(raw):
    if not raw: return raw
    raw = raw.strip()
    m = _LABEL_SPLIT_RE.search(raw)
    return raw[:m.start()].strip() if m and m.start()>0 else raw

def parse_number(text):
    if not text or str(text).strip() in ("-",""): return None
    token = str(text).strip().split()[0].replace(",","")
    try: return int(token)
    except ValueError:
        try: return float(token)
        except: return None

def parse_basic_table(soup):
    data={}; table=soup.select_one(".short-doc table")
    if not table: return data
    basic_map = {
        "CK cơ sở":"ck_co_so","Tổ chức phát hành CKCS":"to_chuc_ph_ckcs",
        "Tổ chức phát hành CW":"to_chuc_ph_cw","Loại chứng quyền":"loai_cw",
        "Kiểu thực hiện":"kieu_thuc_hien","Phương thức thực hiện quyền":"phuong_thuc",
        "Thời hạn":"thoi_han","Ngày phát hành":"ngay_phat_hanh",
        "Ngày niêm yết":"ngay_niem_yet","Ngày giao dịch đầu tiên":"ngay_gd_dau_tien",
        "Ngày giao dịch cuối cùng":"ngay_gd_cuoi_cung","Ngày đáo hạn":"ngay_dao_han",
        "Tỷ lệ chuyển đổi":"ty_le_chuyen_doi","TLCĐ điều chỉnh":"tlcd_dieu_chinh",
        "Giá phát hành":"gia_phat_hanh","Giá thực hiện":"gia_thuc_hien",
        "Giá TH điều chỉnh":"gia_th_dieu_chinh","Khối lượng Niêm yết":"kl_niem_yet",
        "Khối lượng lưu hành":"kl_luu_hanh",
    }
    for tr in table.find_all("tr"):
        tds=tr.find_all("td",limit=2)
        if len(tds)<2: continue
        b=tds[0].find("b")
        label=(b.get_text(strip=True) if b else tds[0].get_text(strip=True)).replace(":","").strip()
        val=clean_cell(tds[1].get_text(" ",strip=True))
        for key,col in basic_map.items():
            if key in label: data[col]=val; break
    return data

def get_cw_detail(code):
    url=f"{BASE_URL}/chung-khoan-phai-sinh/{code}/cw-tong-quan.htm"
    res={"ma_cw":code,"status":"unknown"}
    try: resp=get_session().get(url,timeout=TIMEOUT,allow_redirects=True)
    except requests.RequestException as e:
        res["status"]="error"; res["loi"]=str(e)[:120]; return res
    if resp.status_code==404: res["status"]="not_found"; return res
    if resp.status_code!=200: res["status"]=f"http_{resp.status_code}"; return res
    soup=BeautifulSoup(resp.text,"html.parser")
    if not soup.select_one("h1.h1-title"): res["status"]="not_found"; return res
    res["status"]="found"
    pe=soup.select_one("#stockprice .price")
    res["gia_hien_tai"]=parse_number(pe.get_text(strip=True)) if pe else None
    for sel,key,is_num in [
        ("#basestock","gia_ck_co_so",True),("#moneyness","s_x",True),
        ("#breakeven","hoa_von",True),("#moneyness-status","trang_thai_cw",False)
    ]:
        el=soup.select_one(sel)
        if el:
            v=el.get_text(strip=True)
            res[key]=parse_number(v) if is_num else v
    res.update(parse_basic_table(soup))
    return res

def scrape_codes(codes, label=""):
    total=len(codes); raw={}; lock=threading.Lock(); t0=time.time()
    def crawl(code):
        time.sleep(DELAY); return code,get_cw_detail(code)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs={ex.submit(crawl,c):c for c in codes}
        for f in as_completed(futs):
            code,result=f.result(); raw[code]=result; done=len(raw)
            if result.get("status")=="found":
                el=time.time()-t0; sp=done/el if el else 0; eta=(total-done)/sp if sp else 0
                with lock:
                    print(f"  [{done:>4}/{total}] OK {code}  HH:{result.get('ngay_dao_han',''):12}  {sp:.1f}/s ETA:{eta:.0f}s")
            elif done%100==0:
                el=time.time()-t0; sp=done/el if el else 0; eta=(total-done)/sp if sp else 0
                with lock: print(f"  [{done:>4}/{total}] ... {label}  {sp:.1f}/s ETA:{eta:.0f}s")
    records=[raw[c] for c in codes if raw.get(c,{}).get("status")=="found"]
    df=pd.DataFrame(records) if records else pd.DataFrame()
    if not df.empty:
        df.drop(columns=[c for c in ["status","loi"] if c in df.columns],inplace=True)
    return df

def _refresh_active_prices(active_codes: list, df_cache: pd.DataFrame) -> pd.DataFrame:
    if not active_codes:
        return df_cache
    print(f"   Lightweight price refresh cho {len(active_codes)} CW active...")
    t0 = time.time()
    def fetch_price_only(code):
        url = f"{BASE_URL}/chung-khoan-phai-sinh/{code}/cw-tong-quan.htm"
        try:
            resp = get_session().get(url, timeout=TIMEOUT, allow_redirects=True)
            if resp.status_code != 200:
                return code, None, None
            soup = BeautifulSoup(resp.text, "html.parser")
            pe = soup.select_one("#stockprice .price")
            gia = parse_number(pe.get_text(strip=True)) if pe else None
            el = soup.select_one("#moneyness-status")
            trang_thai = el.get_text(strip=True) if el else None
            return code, gia, trang_thai
        except Exception:
            return code, None, None
    updates = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(fetch_price_only, c): c for c in active_codes}
        for f in as_completed(futs):
            code, gia, trang_thai = f.result()
            updates[code] = (gia, trang_thai)
    df_out = df_cache.copy()
    mask = df_out["ma_cw"].isin(updates)
    for code, (gia, trang_thai) in updates.items():
        idx = df_out.index[df_out["ma_cw"] == code]
        if len(idx) and gia is not None:
            df_out.loc[idx, "gia_hien_tai"] = gia
        if len(idx) and trang_thai is not None:
            df_out.loc[idx, "trang_thai_cw"] = trang_thai
    elapsed = time.time() - t0
    ok = sum(1 for g, _ in updates.values() if g is not None)
    print(f"   Price refresh xong: {ok}/{len(active_codes)} OK | {elapsed:.1f}s")
    return df_out

# ══════════════════════════════════════════════════════════════════
# BUOC 1 - VIETSTOCK (INCREMENTAL)
# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
# BUG FIX B: Auto-discovery CW codes từ Vietstock
# Vấn đề gốc: all_codes dùng template tĩnh MAX_ISSUANCE=50 → bỏ sót
# CW có số thứ tự > 50 hoặc underlying mới chưa có trong BASE_STOCKS.
#
# Giải pháp: Adaptive probing — với mỗi prefix (ví dụ CHPG26),
# probe tiếp lên đến khi gặp N_MISS_TOLERANCE lần 404 liên tiếp.
# Đây là cách duy nhất reliable khi không có public listing API.
# ══════════════════════════════════════════════════════════════════
N_MISS_TOLERANCE = 3   # cho phép N lần 404 liên tiếp trước khi dừng probe

def _probe_prefix(prefix: str, known: set, sess) -> list[str]:
    """
    Với prefix = "CHPG26", probe CHPG2601, CHPG2602, ...
    Dừng khi gặp N_MISS_TOLERANCE lần 404 LIÊN TIẾP.
    Trả về list các mã chưa có trong `known`.
    """
    found = []
    miss_streak = 0
    n = 1
    while True:
        code = f"{prefix}{n:02d}"
        if code in known:
            miss_streak = 0  # mã đã biết → không tính là miss
            n += 1
            continue
        url = f"{BASE_URL}/chung-khoan-phai-sinh/{code}/cw-tong-quan.htm"
        try:
            r = sess.get(url, timeout=TIMEOUT, allow_redirects=True)
        except requests.RequestException:
            miss_streak += 1
            n += 1
            if miss_streak >= N_MISS_TOLERANCE:
                break
            continue
        if r.status_code == 404 or not BeautifulSoup(r.text, "html.parser").select_one("h1.h1-title"):
            miss_streak += 1
            if miss_streak >= N_MISS_TOLERANCE:
                break
        else:
            miss_streak = 0
            found.append(code)
        time.sleep(DELAY)
        n += 1
    return found

def discover_new_cw_codes(known: set) -> list[str]:
    """
    BUG FIX B: Phát hiện CW mới NGOÀI template tĩnh bằng adaptive probing.
    Chạy sau khi đã xử lý new_codes từ template — chỉ probe các prefix
    mà MAX_ISSUANCE của template có thể đã bị vượt qua.
    """
    print("   [Discovery] Adaptive probing cho CW ngoài template tĩnh...")
    sess = get_session()
    all_new = []

    for stock in BASE_STOCKS:
        prefix_base = stock  # vd: "CHPG"
        for yy in YEARS:
            prefix = f"{prefix_base}{yy}"  # vd: "CHPG26"
            # Tìm số thứ tự lớn nhất đã biết cho prefix này
            existing_nums = [
                int(c.replace(prefix, ""))
                for c in known
                if c.startswith(prefix) and c.replace(prefix, "").isdigit()
            ]
            max_known = max(existing_nums) if existing_nums else 0

            # Chỉ probe tiếp nếu đã có CW gần MAX_ISSUANCE (ngưỡng 80%)
            # hoặc nếu đây là năm hiện tại (26) → luôn probe
            current_yy = str((datetime.now(timezone.utc) + timedelta(hours=7)).year)[-2:]
            is_current_year = (yy == current_yy)
            near_limit = max_known >= int(MAX_ISSUANCE * 0.8)

            if not (is_current_year or near_limit):
                continue  # prefix này không cần probe thêm

            # Probe từ max_known+1 trở đi
            probe_start = max(max_known + 1, MAX_ISSUANCE + 1)
            miss_streak = 0
            n = probe_start
            while True:
                code = f"{prefix}{n:02d}"
                if code in known:
                    miss_streak = 0
                    n += 1
                    continue
                url = f"{BASE_URL}/chung-khoan-phai-sinh/{code}/cw-tong-quan.htm"
                try:
                    r = sess.get(url, timeout=TIMEOUT, allow_redirects=True)
                    soup = BeautifulSoup(r.text, "html.parser")
                    if r.status_code != 200 or not soup.select_one("h1.h1-title"):
                        miss_streak += 1
                    else:
                        miss_streak = 0
                        all_new.append(code)
                        print(f"   [Discovery] Tìm thấy mã mới ngoài template: {code}")
                except requests.RequestException:
                    miss_streak += 1
                time.sleep(DELAY)
                n += 1
                if miss_streak >= N_MISS_TOLERANCE:
                    break

    if all_new:
        print(f"   [Discovery] Tổng cộng {len(all_new)} mã mới ngoài template: {all_new}")
    else:
        print("   [Discovery] Không tìm thấy mã nào ngoài template.")
    return all_new

def step1_vietstock():
    print("\n"+"="*60)
    print("BUOC 1 - Vietstock (incremental + lightweight daily refresh)")
    print("="*60)

    # Template tĩnh — baseline đã biết
    all_codes = [
        f"{s}{yy}{n:02d}"
        for s in BASE_STOCKS for yy in YEARS for n in range(1, MAX_ISSUANCE + 1)
    ]

    df_cache = load_cache(VIETSTOCK_CACHE)
    if df_cache is None:
        print(f"   Full load - {len(all_codes):,} ma tu template")
        df = scrape_codes(all_codes, "full")
        save_cache(df, VIETSTOCK_CACHE)
        return df

    known = set(df_cache["ma_cw"].tolist())

    # ── Bước 1a: Mã mới trong template chưa có trong cache ──
    new_codes = [c for c in all_codes if c not in known]
    print(f"   Cache co: {len(known)} ma  |  Ma moi (template): {len(new_codes)}")
    if new_codes:
        df_new = scrape_codes(new_codes, "new CW (template)")
        if not df_new.empty:
            df_cache = pd.concat([df_cache, df_new], ignore_index=True)
            df_cache.drop_duplicates(subset="ma_cw", keep="last", inplace=True)
            known = set(df_cache["ma_cw"].tolist())
            print(f"   Them {len(df_new)} ma moi (template) vao cache")

    # ── Bước 1b: BUG FIX B — Adaptive discovery ngoài template ──
    # Phát hiện CW có số thứ tự > MAX_ISSUANCE hoặc underlying mới
    extra_codes = discover_new_cw_codes(known)
    if extra_codes:
        df_extra = scrape_codes(extra_codes, "new CW (discovery)")
        if not df_extra.empty:
            df_cache = pd.concat([df_cache, df_extra], ignore_index=True)
            df_cache.drop_duplicates(subset="ma_cw", keep="last", inplace=True)
            print(f"   Them {len(df_extra)} ma moi (discovery) vao cache → tong: {len(df_cache)}")

    # ── Bước 1c: BUG FIX C — Dùng ICT, không dùng date.today() UTC ──
    today_ict = (datetime.now(timezone.utc) + timedelta(hours=7)).date()
    today_ts  = pd.Timestamp(today_ict)
    if "ngay_gd_cuoi_cung" in df_cache.columns:
        ldt = pd.to_datetime(df_cache["ngay_gd_cuoi_cung"], dayfirst=True, errors="coerce")
        active_codes = df_cache.loc[ldt >= today_ts, "ma_cw"].tolist()
    else:
        active_codes = []
    df_cache = _refresh_active_prices(active_codes, df_cache)
    save_cache(df_cache, VIETSTOCK_CACHE)
    return df_cache

# ══════════════════════════════════════════════════════════════════
# BUOC 2 - OHLCV (INCREMENTAL) - vnstock
# ══════════════════════════════════════════════════════════════════
def _to_ymd(s):
    s = str(s).strip()
    if re.match(r'\d{2}/\d{2}/\d{4}', s):
        d,m,y = s.split("/"); return f"{y}-{m}-{d}"
    return s

def _normalise_ohlcv(df, symbol):
    df = df.copy()
    df.columns = [c.lower() for c in df.columns]
    for alt in ["time","date","trading_date","datetime"]:
        if alt in df.columns:
            if alt != "time":
                df.rename(columns={alt:"time"}, inplace=True)
            break
    for alt in ["vol","klgd"]:
        if alt in df.columns:
            df.rename(columns={alt:"volume"}, inplace=True); break
    for src in ["open","high","low","close"]:
        if src not in df.columns:
            for alt in [f"{src}_price", f"gia_{src}"]:
                if alt in df.columns:
                    df.rename(columns={alt:src}, inplace=True); break
    if "time" in df.columns:
        col = df["time"].astype(str).str.strip()
        col = col.str.replace(r"\s+\d{2}:\d{2}.*$", "", regex=True)
        parsed = pd.to_datetime(col, errors="coerce")
        if parsed.isna().sum() > len(df) * 0.3:
            parsed = pd.to_datetime(col, dayfirst=True, errors="coerce")
        df["time"] = parsed.dt.strftime("%d/%m/%Y")
    df["Ticker"] = symbol
    keep = [c for c in ["time","open","high","low","close","volume","Ticker"] if c in df.columns]
    return df[keep].dropna(subset=["time"])

def fetch_one(symbol, start_str, end_str, _vci_circuit: dict | None = None):
    """
    Fetch OHLCV cho 1 CW/underlying.

    FIX 1: Trigger KBS fallback khi ReadTimeout (không chỉ 403)
    FIX 2: Giới hạn read_timeout trong vnstock SDK call xuống 20s (thay vì 30s mặc định)
    FIX 3: Circuit breaker — nếu VCI liên tiếp timeout >= CIRCUIT_OPEN_AFTER lần,
           tự động skip VCI và dùng KBS cho các lần gọi tiếp trong cùng 1 run
    """
    CIRCUIT_OPEN_AFTER = 6   # số lần timeout liên tiếp toàn cục để mở circuit
    VCI_READ_TIMEOUT   = 20  # giây — giảm từ 30s mặc định của SDK

    start = _to_ymd(start_str)
    end   = _to_ymd(end_str)

    # _vci_circuit là dict mutable dùng chung toàn bộ run: {"fails": N, "open": bool}
    if _vci_circuit is None:
        _vci_circuit = fetch_one._circuit  # dùng state cấp module

    def _try_vci():
        from vnstock.api.quote import Quote
        import requests as _req
        # FIX 2: monkey-patch session timeout của vnstock nếu có thể,
        # nếu không thì dùng threading.Timer để hard-cap
        q = Quote(symbol=symbol, source="VCI")
        # Thử set timeout trực tiếp trên session nếu SDK expose nó
        if hasattr(q, 'session') and hasattr(q.session, 'request'):
            q.session.request = lambda method, url, **kw: (
                kw.update({"timeout": (10, VCI_READ_TIMEOUT)}) or
                type(q.session).request(q.session, method, url, **kw)
            )
        df = q.history(start=start, end=end, interval="1D")
        return df

    def _try_kbs():
        from vnstock.api.quote import Quote as Q2
        q2 = Q2(symbol=symbol, source="KBS")
        df2 = q2.history(start=start, end=end, interval="1D")
        return df2

    def _normalise(df):
        if df is None or df.empty:
            return pd.DataFrame()
        return _normalise_ohlcv(df, symbol)

    # FIX 3: Nếu circuit đang open → bỏ qua VCI luôn, thử KBS trực tiếp
    if _vci_circuit.get("open"):
        try:
            df = _try_kbs()
            result = _normalise(df)
            if not result.empty:
                return result
        except Exception:
            pass
        return pd.DataFrame()

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            df = _try_vci()
            # Thành công → reset fail streak
            _vci_circuit["fails"] = 0
            if df is None or df.empty:
                return pd.DataFrame()
            return _normalise(df)

        except Exception as e:
            msg = str(e)
            is_timeout = any(k in msg.lower() for k in
                             ["timed out", "timeout", "read timeout", "connect timeout",
                              "connectionerror", "remotedisconnected"])
            is_rl  = any(k in msg.lower() for k in
                         ["rate limit","429","too many","quota","throttle",
                          "gian han api","gioi han","rate_limit","exceeded"])
            is_403 = "403" in msg or "forbidden" in msg.lower()

            if is_rl:
                wait = 60
                print(f"      RL {symbol} #{attempt} → cho {wait}s roi thu lai...")
                time.sleep(wait)
                continue

            # FIX 1: ReadTimeout → tăng fail streak, thử KBS ngay thay vì retry VCI
            if is_timeout:
                _vci_circuit["fails"] = _vci_circuit.get("fails", 0) + 1
                print(f"      TIMEOUT {symbol} #{attempt} | VCI streak={_vci_circuit['fails']}")

                # FIX 3: Mở circuit nếu đủ ngưỡng
                if _vci_circuit["fails"] >= CIRCUIT_OPEN_AFTER:
                    _vci_circuit["open"] = True
                    print(f"      [CIRCUIT OPEN] VCI unstable ({_vci_circuit['fails']} timeouts) "
                          f"→ chuyển sang KBS cho toàn bộ run")

                # FIX 1: Thử KBS ngay thay vì retry VCI tiếp
                try:
                    df2 = _try_kbs()
                    result = _normalise(df2)
                    if not result.empty:
                        print(f"      KBS fallback OK {symbol}")
                        return result
                except Exception as e2:
                    print(f"      KBS fallback FAIL {symbol}: {str(e2)[:50]}")

                # KBS cũng fail → nếu còn attempt thì retry VCI, ngược lại bỏ
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                continue

            # 403 hoặc lỗi khác → thử KBS ở attempt đầu
            if is_403 and attempt == 1:
                try:
                    df2 = _try_kbs()
                    result = _normalise(df2)
                    if not result.empty:
                        return result
                except Exception:
                    pass

            tag = "403" if is_403 else "ERR"
            print(f"      {tag} {symbol} #{attempt} wait {RETRY_DELAY:.0f}s | {msg[:60]}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)

    return None

# Circuit breaker state — dict mutable dùng chung cho cả run (reset khi module load lại)
fetch_one._circuit = {"fails": 0, "open": False}

def _is_trading_day_closed() -> bool:
    now_ict = datetime.now(timezone.utc) + timedelta(hours=7)
    if now_ict.weekday() >= 5:
        return True
    if now_ict.hour > 15 or (now_ict.hour == 15 and now_ict.minute >= 30):
        return True
    return False

def step2_ohlcv(df_vs):
    print("\n"+"="*60)
    print("BUOC 2 - OHLCV (incremental)")
    print("="*60)
    api_key = os.environ.get("VNSTOCK_API","")
    if api_key:
        try:
            import vnai; vnai.setup_api_key(api_key)
            print("   OK Vnstock API key registered.")
        except Exception as e:
            print(f"   WARN: setup_api_key failed: {e}")
    else:
        print("   WARN: no VNSTOCK_API env var.")
    tickers = df_vs["ma_cw"].dropna().unique().tolist()
    df_cache = load_cache(OHLCV_CACHE)
    now_ict = datetime.now(timezone.utc) + timedelta(hours=7)
    today = now_ict.date()
    today_str = today.strftime("%Y-%m-%d")
    end_str = (today + timedelta(days=1)).strftime("%Y-%m-%d")
    if df_cache is None:
        print(f"   Full load - {len(tickers)} ma tu {OHLCV_START_DATE}")
        print(f"   end_str = {end_str} (today+1, VCI exclusive)")
        rows=[]; failed=[]; skipped=[]
        fetch_one._circuit = {"fails": 0, "open": False}  # reset circuit cho run mới
        for i,sym in enumerate(tickers,1):
            df_r = fetch_one(sym, OHLCV_START_DATE, end_str, fetch_one._circuit)
            if df_r is None:
                failed.append(sym)
                print(f"  [{i:>4}/{len(tickers)}] FAIL {sym}")
            elif df_r.empty:
                skipped.append(sym)
            else:
                rows.append(df_r)
                if i % 50 == 1:
                    print(f"  [{i:>4}/{len(tickers)}] OK {sym} ({len(df_r)} phien)")
            time.sleep(REQUEST_DELAY)
        df_out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
        print(f"   OK:{len(rows)}  Empty:{len(skipped)}  Fail:{len(failed)}")
        if failed: print(f"   Failed: {failed}")
        save_cache(df_out, OHLCV_CACHE)
        return df_out
    df_cache["time"] = df_cache["time"].astype(str)
    df_cache["time_dt"] = pd.to_datetime(df_cache["time"], dayfirst=True, errors="coerce")
    last_dt = df_cache.groupby("Ticker")["time_dt"].max()
    cached = set(last_dt.index)
    is_closed = _is_trading_day_closed()
    last_expected_session = today if is_closed else (today - timedelta(days=1))
    while last_expected_session.weekday() >= 5:
        last_expected_session -= timedelta(days=1)
    print(f"   Ngay hien tai (ICT): {today}  |  Gio ICT: {now_ict.strftime('%H:%M')}")
    print(f"   end_str fetch       : {end_str} (today+1, VCI exclusive)")
    print(f"   Gio ICT: {now_ict.strftime('%H:%M')} | HoSE {'DA DONG' if is_closed else 'CHUA DONG'}")
    print(f"   Phien can co trong cache: {last_expected_session}")
    # ── Chuẩn bị lookup ngay_gd_dau_tien từ Vietstock metadata ──
    # BUG FIX A: Dùng ngay_gd_dau_tien làm start_str cho CW mới thay vì OHLCV_START_DATE
    # → Tránh fetch thừa từ 2023 với CW niêm yết năm 2026, đặc biệt quan trọng
    #   khi vnstock chưa index kịp phiên đầu tiên → lần fetch tiếp theo sẽ biết đúng điểm bắt đầu
    first_trading_map = {}
    if "ngay_gd_dau_tien" in df_vs.columns:
        for _, row in df_vs[["ma_cw", "ngay_gd_dau_tien"]].dropna().iterrows():
            raw = str(row["ngay_gd_dau_tien"]).strip()
            try:
                dt = pd.to_datetime(raw, dayfirst=True, errors="coerce")
                if pd.notna(dt):
                    # Lùi thêm 1 ngày để buffer an toàn (tránh bỏ sót do off-by-one)
                    first_trading_map[row["ma_cw"]] = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
            except Exception:
                pass

    to_fetch = []
    skipped_count = 0
    new_count = 0
    backfill_count = 0

    for sym in tickers:
        # Xác định start_str đúng cho CW này (ưu tiên ngay_gd_dau_tien)
        known_start = first_trading_map.get(sym, OHLCV_START_DATE)

        if sym in cached:
            last = last_dt[sym]
            if pd.isna(last):
                start_str = (today - timedelta(days=30)).strftime("%Y-%m-%d")
                to_fetch.append((sym, start_str, "NaT-reload"))

            elif last.date() >= last_expected_session:
                # Cache đã đủ phiên mới nhất — nhưng kiểm tra back-fill:
                # BUG FIX A CORE: CW mới có thể bị thiếu phiên đầu tiên
                # vì lần fetch đầu trả về empty (vnstock chưa index kịp).
                # Kiểm tra: first_trading_map[sym] < first session trong cache
                first_in_cache = df_cache.loc[df_cache["Ticker"] == sym, "time_dt"].min()
                if pd.notna(first_in_cache) and sym in first_trading_map:
                    expected_start = pd.to_datetime(first_trading_map[sym])
                    # Nếu phiên đầu trong cache muộn hơn ngay_gd_dau_tien + 2 ngày → cần back-fill
                    gap_days = (first_in_cache - expected_start).days
                    if gap_days > 2:
                        backfill_start = first_trading_map[sym]
                        backfill_end = (first_in_cache - timedelta(days=1)).strftime("%Y-%m-%d")
                        to_fetch.append((sym, backfill_start, f"backfill-{backfill_start}→{backfill_end}"))
                        backfill_count += 1
                        continue
                skipped_count += 1

            else:
                next_day = (last + timedelta(days=1)).strftime("%Y-%m-%d")
                to_fetch.append((sym, next_day, f"+tu {(last+timedelta(days=1)).strftime('%d/%m/%Y')}"))
        else:
            # CW hoàn toàn mới: dùng ngay_gd_dau_tien thay vì OHLCV_START_DATE (2023-01-01)
            to_fetch.append((sym, known_start, f"new (tu {known_start})"))
            new_count += 1

    print(f"   Skip (cache du)   : {skipped_count} ma")
    print(f"   Can fetch moi     : {new_count} ma")
    print(f"   Back-fill phien   : {backfill_count} ma  ← BUG FIX A")
    print(f"   Can update phien  : {len(to_fetch) - new_count - backfill_count} ma")
    print(f"   Tong can fetch    : {len(to_fetch)} ma | delay {REQUEST_DELAY}s/ma")
    if not to_fetch:
        print("   → Cache day du, SKIP toan bo OHLCV fetch.")
        df_cache.drop(columns=["time_dt"], inplace=True)
        return df_cache
    print(f"   ETA               : ~{len(to_fetch) * REQUEST_DELAY / 60:.1f} phut")
    DELAY_OK    = REQUEST_DELAY
    DELAY_EMPTY = 0.15
    DELAY_FAIL  = 0.3
    new_rows = []; backfill_rows = []; failed = []; n_ok = 0; n_empty = 0
    t_fetch_start = time.time()
    fetch_one._circuit = {"fails": 0, "open": False}  # reset circuit cho run mới
    for i, (sym, start_str, lbl) in enumerate(to_fetch, 1):
        df_r = fetch_one(sym, start_str, end_str, fetch_one._circuit)
        if df_r is None:
            failed.append(sym)
            print(f"  [{i:>4}/{len(to_fetch)}] FAIL {sym}")
            time.sleep(DELAY_FAIL)
        elif df_r.empty:
            n_empty += 1
            if "new" in lbl or "NaT" in lbl or "backfill" in lbl:
                print(f"  [{i:>4}/{len(to_fetch)}] NO_NEW {sym} ({lbl})")
            time.sleep(DELAY_EMPTY)
        else:
            # BUG FIX A: back-fill rows được xử lý riêng để prepend (thêm vào đầu)
            if "backfill" in lbl:
                backfill_rows.append(df_r)
                print(f"  [{i:>4}/{len(to_fetch)}] BACKFILL {sym} +{len(df_r)} phien ({lbl})")
            else:
                new_rows.append(df_r)
            n_ok += 1
            if n_ok % 50 == 1 or "new" in lbl or "NaT" in lbl:
                elapsed = time.time() - t_fetch_start
                eta = (len(to_fetch) - i) * (elapsed / i) / 60
                print(f"  [{i:>4}/{len(to_fetch)}] OK {sym} +{len(df_r)} ({lbl}) | OK={n_ok} ETA={eta:.1f}ph")
            time.sleep(DELAY_OK)
    df_cache.drop(columns=["time_dt"], inplace=True)
    print(f"\n   Ket qua: OK={n_ok} | NO_NEW={n_empty} | FAIL={len(failed)} | SKIP={skipped_count} | BACKFILL={len(backfill_rows)}")
    if failed: print(f"   Failed: {failed[:10]}")
    all_new_rows = backfill_rows + new_rows   # BUG FIX A: gộp back-fill + phiên mới
    if all_new_rows:
        df_add = pd.concat(all_new_rows, ignore_index=True)
        bad_mask = df_cache["time"].isna() | (df_cache["time"].astype(str) == "NaT")
        n_bad = bad_mask.sum()
        if n_bad > 0:
            print(f"   Xoa {n_bad} rows NaT trong cache")
            df_cache = df_cache[~bad_mask].copy()
        df_merged = pd.concat([df_cache, df_add], ignore_index=True)
        df_merged["time"] = df_merged["time"].astype(str).str.strip()
        df_merged = df_merged[~df_merged["time"].isin(["NaT","nan",""])]
        df_merged["_sort_dt"] = pd.to_datetime(df_merged["time"], dayfirst=True, errors="coerce")
        df_merged.sort_values(["Ticker","_sort_dt"], inplace=True)
        df_merged.drop_duplicates(subset=["time","Ticker"], keep="last", inplace=True)
        df_merged.drop(columns=["_sort_dt"], inplace=True)
        df_merged.reset_index(drop=True, inplace=True)
        print(f"   Them {len(df_add):,} dong moi → tong {len(df_merged):,} dong")
        save_cache(df_merged, OHLCV_CACHE)
        return df_merged
    else:
        print("   Khong co du lieu moi.")
        if failed: print(f"   Failed: {failed}")
        save_cache(df_cache, OHLCV_CACHE)
        return df_cache

# ══════════════════════════════════════════════════════════════════
# BUOC 3 - LOC + SORT
# ══════════════════════════════════════════════════════════════════
def step3_filter(df_full):
    print("\n"+"="*60)
    print("BUOC 3 - Loc & sort")
    print("="*60)
    df_full = df_full.copy()
    df_full["time"] = df_full["time"].astype(str).str.strip()
    df_full = df_full[~df_full["time"].isin(["NaT","nan","","None"])].copy()
    df_full["time_dt"] = pd.to_datetime(df_full["time"], dayfirst=True, errors="coerce")
    df_full = df_full[df_full["time_dt"].notna()].copy()
    ltd = df_full.groupby("Ticker")["time_dt"].max().reset_index().rename(columns={"time_dt":"ltd"})
    ts = pd.Timestamp(FILTER_DATE)
    valid = ltd.loc[ltd["ltd"]>=ts,"Ticker"].tolist()
    removed = ltd.loc[ltd["ltd"]<ts,"Ticker"].tolist()
    print(f"   Tong ma: {len(ltd)}  |  Hop le: {len(valid)}  |  Loai: {len(removed)}")
    df = df_full[df_full["Ticker"].isin(valid) & (df_full["time_dt"]>=ts)].copy()
    df.sort_values(["time_dt","Ticker"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    df["time"] = df["time_dt"].dt.strftime("%d/%m/%Y")
    df.drop(columns=["time_dt"], inplace=True)
    cols = [c for c in ["time","open","high","low","close","volume","Ticker"] if c in df.columns]
    print(f"   Phien GD: {len(df):,}  |  Ngay: {df['time'].nunique()}  |  CW: {df['Ticker'].nunique()}")
    return df[cols], valid

# ══════════════════════════════════════════════════════════════════
# BUOC 4 - XUAT EXCEL
# ══════════════════════════════════════════════════════════════════
COL_RENAME={
    "ma_cw":"Ticker","ck_co_so":"Underlying Asset","to_chuc_ph_cw":"Issuer",
    "loai_cw":"Type","kieu_thuc_hien":"Exercise Style","thoi_han":"Term",
    "ngay_phat_hanh":"Issuance Date","ngay_niem_yet":"Listing Date",
    "ngay_gd_dau_tien":"First Trading Date","ngay_gd_cuoi_cung":"Last Trading Date",
    "ngay_dao_han":"Maturity Date","ty_le_chuyen_doi":"Conversion Ratio",
    "tlcd_dieu_chinh":"Adj. Conversion Ratio","gia_thuc_hien":"Exercise Price",
    "gia_th_dieu_chinh":"Adj. Exercise Price","gia_phat_hanh":"Issuance Price",
    "kl_niem_yet":"Listed Volume","kl_luu_hanh":"Outstanding Volume",
    "gia_hien_tai":"Current Price","trang_thai_cw":"Moneyness","s_x":"S/X","hoa_von":"Break-even",
}
PRIORITY=["ma_cw","ck_co_so","to_chuc_ph_cw","ngay_gd_dau_tien","ngay_gd_cuoi_cung",
          "ngay_dao_han","ty_le_chuyen_doi","tlcd_dieu_chinh","gia_thuc_hien",
          "gia_th_dieu_chinh","kl_niem_yet","kl_luu_hanh","loai_cw","kieu_thuc_hien",
          "thoi_han","ngay_phat_hanh","ngay_niem_yet","gia_phat_hanh","gia_hien_tai","trang_thai_cw"]

def fmt_sheet(ws, color):
    fill=PatternFill("solid",fgColor=color); font=Font(bold=True,color="FFFFFF",size=10)
    aln=Alignment(horizontal="center",vertical="center")
    brd=Border(left=Side(style="thin"),right=Side(style="thin"),
               top=Side(style="thin"),bottom=Side(style="thin"))
    for c in ws[1]: c.fill=fill; c.font=font; c.alignment=aln; c.border=brd
    ws.freeze_panes="A2"
    for col in ws.columns:
        w=max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(w+2,30)

def step4_excel(df_ohlcv, df_vs, valid_tickers):
    print("\n"+"="*60)
    print("BUOC 4 - Xuat Excel")
    print("="*60)
    df=df_vs[df_vs["ma_cw"].isin(valid_tickers)].copy()
    p=[c for c in PRIORITY if c in df.columns]
    df=df[p+[c for c in df.columns if c not in p]]
    today_ts = pd.Timestamp((datetime.now(timezone.utc) + timedelta(hours=7)).date())  # BUG FIX C
    df["_ldt"]=pd.to_datetime(df["ngay_gd_cuoi_cung"],dayfirst=True,errors="coerce")
    df.rename(columns=COL_RENAME,inplace=True)
    act=df[df["_ldt"]>=today_ts].drop(columns=["_ldt"]).sort_values("Ticker")
    exp=df[df["_ldt"]< today_ts].drop(columns=["_ldt"]).sort_values("Ticker")
    os.makedirs("output",exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE,engine="openpyxl") as w:
        df_ohlcv.to_excel(w,sheet_name="OHLCV",index=False)
        act.to_excel(w,sheet_name="CW_Info_Active",index=False)
        exp.to_excel(w,sheet_name="CW_Info_Expired",index=False)
        wb=w.book
        fmt_sheet(wb["OHLCV"],"1F4E79")
        fmt_sheet(wb["CW_Info_Active"],"375623")
        fmt_sheet(wb["CW_Info_Expired"],"843C0C")
    print(f"OK Saved: {OUTPUT_FILE}")
    print(f"   Updated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"   OHLCV: {len(df_ohlcv):,} rows | Active: {len(act)} | Expired: {len(exp)}")

# ══════════════════════════════════════════════════════════════════
# BUOC 5 - XUAT data.json (ĐÃ TÍCH HỢP BLACK‑SCHOLES & LỢI SUẤT HNX)
# ══════════════════════════════════════════════════════════════════
def _parse_ratio(raw):
    try: return float(str(raw).split(":")[0].replace(",","."))
    except: return 1.0

def _parse_exercise(raw):
    try:
        s = re.sub(r"[^\d]","", str(raw))
        return int(s) if s else 0
    except: return 0

def _fetch_underlying_prices(underlyings: list) -> dict:
    prices = {}
    today_str = (datetime.now(timezone.utc) + timedelta(hours=7)).strftime("%Y-%m-%d")  # BUG FIX C
    week_ago = ((datetime.now(timezone.utc) + timedelta(hours=7)).date() - timedelta(days=10)).strftime("%Y-%m-%d")  # BUG FIX C
    for sym in underlyings:
        for source in ["VCI", "KBS"]:
            try:
                from vnstock.api.quote import Quote
                df = Quote(symbol=sym, source=source).history(
                    start=week_ago, end=today_str, interval="1D"
                )
                if df is None or df.empty:
                    continue
                df.columns = [c.lower() for c in df.columns]
                close_col = next((c for c in ["close","close_price"] if c in df.columns), None)
                if close_col is None:
                    continue
                last_close = float(df[close_col].iloc[-1])
                if last_close < 1000:
                    last_close *= 1000
                prices[sym] = round(last_close)
                print(f"   underlying {sym}: {last_close:,.0f} VND (source={source})")
                break
            except Exception as e:
                print(f"   WARN underlying {sym} source={source}: {str(e)[:60]}")
                time.sleep(0.5)
                continue
        if sym not in prices:
            print(f"   WARN: khong lay duoc gia {sym}, dung fallback 0")
            prices[sym] = 0
    return prices

def step5_export_json(df_ohlcv_full, df_vietstock, valid_tickers):
    print("\n"+"="*60)
    print("BUOC 5 - Xuat data.json (bao gồm Black‑Scholes)")
    print("="*60)

    today_ts = pd.Timestamp((datetime.now(timezone.utc) + timedelta(hours=7)).date())  # BUG FIX C
    df_info = df_vietstock[df_vietstock["ma_cw"].isin(valid_tickers)].copy()
    df_info["_last_gd_dt"] = pd.to_datetime(df_info["ngay_gd_cuoi_cung"], dayfirst=True, errors="coerce")

    # --- Lấy giá cổ phiếu cơ sở mới nhất ---
    underlyings = df_info["ck_co_so"].dropna().unique().tolist()
    underlyings = [u.strip() for u in underlyings if u.strip()]
    print(f"   Lay gia {len(underlyings)} ma underlying: {underlyings}")
    underlying_prices = _fetch_underlying_prices(underlyings)

    # --- Chuẩn bị dữ liệu OHLCV cho CW (đã có) và underlying (có thể cache riêng) ---
    df_ohlcv_full = df_ohlcv_full.copy()
    df_ohlcv_full["time"] = df_ohlcv_full["time"].astype(str).str.strip()
    df_ohlcv_full = df_ohlcv_full[~df_ohlcv_full["time"].isin(["NaT","nan","","None"])].copy()
    df_ohlcv_full["time_dt"] = pd.to_datetime(df_ohlcv_full["time"], dayfirst=True, errors="coerce")
    df_ohlcv_full = df_ohlcv_full[df_ohlcv_full["time_dt"].notna()].copy()
    ohlcv_idx = {t: grp.set_index("time_dt") for t, grp in df_ohlcv_full.groupby("Ticker")}

    # --- Lấy OHLCV underlying (để tính sigma và premium daily) ---
    UNDERLYING_CACHE = f"{CACHE_DIR}/underlying.parquet"
    underlying_ohlcv = {}
    now_ict = datetime.now(timezone.utc) + timedelta(hours=7)
    today = now_ict.date()
    today_str = today.strftime("%Y-%m-%d")
    end_str = (today + timedelta(days=1)).strftime("%Y-%m-%d")
    _last_exp_und = today if _is_trading_day_closed() else (today - timedelta(days=1))
    while _last_exp_und.weekday() >= 5:
        _last_exp_und -= timedelta(days=1)

    df_und_cache = load_cache(UNDERLYING_CACHE) if os.path.exists(UNDERLYING_CACHE) else None
    und_cache_idx = {}
    if df_und_cache is not None and not df_und_cache.empty:
        df_und_cache["time"] = df_und_cache["time"].astype(str)
        df_und_cache["time_dt"] = pd.to_datetime(df_und_cache["time"], dayfirst=True, errors="coerce")
        und_cache_idx = {t: grp.set_index("time_dt") for t, grp in df_und_cache.groupby("Ticker")}

    und_new_rows = []
    for sym in underlyings:
        if sym in ohlcv_idx:
            underlying_ohlcv[sym] = ohlcv_idx[sym]
            continue
        if sym in und_cache_idx:
            cached_und = und_cache_idx[sym]
            last_und = cached_und.index.max()
            if last_und.date() >= _last_exp_und:
                underlying_ohlcv[sym] = cached_und
                continue
            next_day = (last_und + timedelta(days=1)).strftime("%Y-%m-%d")
            print(f"   underlying {sym}: incremental {next_day} → {end_str}")
            df_u_new = fetch_one(sym, next_day, end_str, fetch_one._circuit)
            if df_u_new is not None and not df_u_new.empty:
                df_u_new["time_dt"] = pd.to_datetime(df_u_new["time"], dayfirst=True, errors="coerce")
                df_u_new = df_u_new[df_u_new["time_dt"].notna()].copy()
                und_new_rows.append(df_u_new)
                combined = pd.concat([cached_und.reset_index(), df_u_new]).drop_duplicates(
                    subset=["time"], keep="last"
                ).set_index("time_dt").sort_index()
                underlying_ohlcv[sym] = combined
                print(f"   underlying {sym}: +{len(df_u_new)} phien → {len(combined)} tong")
            else:
                underlying_ohlcv[sym] = cached_und
            time.sleep(0.4)
            continue
        print(f"   underlying {sym}: full fetch tu {OHLCV_START_DATE} → {end_str}")
        df_u = fetch_one(sym, OHLCV_START_DATE, end_str, fetch_one._circuit)
        if df_u is not None and not df_u.empty:
            df_u["time_dt"] = pd.to_datetime(df_u["time"], dayfirst=True, errors="coerce")
            df_u = df_u[df_u["time_dt"].notna()].copy()
            und_new_rows.append(df_u)
            underlying_ohlcv[sym] = df_u.set_index("time_dt")
            print(f"   underlying {sym}: fetch OK ({len(df_u)} phien)")
        else:
            underlying_ohlcv[sym] = pd.DataFrame()
            print(f"   underlying {sym}: khong lay duoc")
        time.sleep(0.4)

    if und_new_rows:
        df_und_new = pd.concat(und_new_rows, ignore_index=True)
        if df_und_cache is not None and not df_und_cache.empty:
            df_und_cache.drop(columns=["time_dt"], errors="ignore", inplace=True)
            df_und_merged = pd.concat([df_und_cache, df_und_new], ignore_index=True)
            df_und_merged["time"] = df_und_merged["time"].astype(str)
            df_und_merged.drop_duplicates(subset=["time","Ticker"], keep="last", inplace=True)
        else:
            df_und_merged = df_und_new
        save_cache(df_und_merged, UNDERLYING_CACHE)
        print(f"   underlying cache updated: {len(df_und_merged)} dong")

    # --- Lấy lãi suất phi rủi ro (10Y) ---
    bond_df = fetch_bond_yields(start_date="2016-01-01", end_date=today_str)
    if bond_df.empty or "10Y" not in bond_df.columns:
        print("[!] Không có dữ liệu lợi suất, dùng fallback 4.53%")
        risk_free = 0.0453
    else:
        latest_yield = bond_df["10Y"].dropna().iloc[-1]
        risk_free = latest_yield / 100.0
        print(f"[+] Lãi suất phi rủi ro (10Y): {risk_free:.4f}")

    # --- Hàm lấy chuỗi giá cơ sở cho dashboard ---
    def get_underlying_price_series(sym):
        df_u = underlying_ohlcv.get(sym, pd.DataFrame())
        if df_u.empty or "close" not in df_u.columns:
            return {}
        result = {}
        for dt, row in df_u.iterrows():
            try:
                price = float(row["close"])
                if price < 1000:
                    price *= 1000
                result[dt.strftime("%d/%m/%Y")] = round(price)
            except Exception:
                continue
        return result
    underlying_price_series = {sym: get_underlying_price_series(sym) for sym in underlyings}

    # --- Xây dựng danh sách CW với các chỉ số BS ---
    cw_list = []
    for _, row in df_info.iterrows():
        ticker = row["ma_cw"]
        und_sym = str(row.get("ck_co_so", "")).strip()
        ratio = _parse_ratio(row.get("ty_le_chuyen_doi", 1))
        exercise = _parse_exercise(row.get("gia_thuc_hien", 0))  # VND
        loai_cw = str(row.get("loai_cw", "")).lower()
        is_call = True if "mua" in loai_cw else ("bán" not in loai_cw)  # default call

        # Giá hiện tại của CW và cổ phiếu cơ sở
        sub = ohlcv_idx.get(ticker, pd.DataFrame())
        price_cw = float(sub["close"].iloc[-1]) if not sub.empty else 0.0
        S_val = underlying_prices.get(und_sym, 0)

        # Tính premium, đòn bẩy (giữ lại để hiển thị)
        if S_val > 0 and price_cw > 0 and ratio > 0 and exercise > 0:
            price_cw_vnd = price_cw * 1000
            intrinsic = max(S_val - exercise, 0)
            cw_value = price_cw_vnd * ratio
            premium = (cw_value - intrinsic) / S_val * 100
            gross_lev = S_val / cw_value if cw_value > 0 else 0
            moneyness = "ITM" if S_val > exercise * 1.02 else "OTM" if S_val < exercise * 0.98 else "ATM"
            breakeven = (exercise + cw_value) / 1000
        else:
            premium = gross_lev = 0.0
            moneyness = "N/A"
            breakeven = 0.0

        # --- Black‑Scholes cho Call và Put ---
        # Thời gian đến đáo hạn (năm)
        mat_date = row.get("ngay_dao_han")
        if mat_date:
            try:
                maturity = pd.to_datetime(mat_date, dayfirst=True)
                T = (maturity - pd.Timestamp((datetime.now(timezone.utc) + timedelta(hours=7)).date())).days / 365.0  # BUG FIX C
            except:
                T = 0.0
        else:
            T = 0.0

        # Sigma từ lịch sử giá cổ phiếu cơ sở (252 phiên)
        sigma = 0.0
        if und_sym in underlying_ohlcv and not underlying_ohlcv[und_sym].empty:
            close_prices = underlying_ohlcv[und_sym]["close"].values
            sigma = historical_volatility(close_prices, window=252)
        if sigma <= 0:
            sigma = 0.25   # fallback nếu không đủ dữ liệu

        bs_call = bs_put = {"bs_price": 0, "delta_bs": 0, "gamma": 0, "vega": 0, "theta": 0}

        # BUG FIX ordering: tính exercise_vnd TRƯỚC khi gọi BS (không phải sau!)
        # exercise từ parse_exercise() có thể là nghìn đồng → chuẩn hóa về VND gốc
        exercise_vnd = exercise
        if exercise > 0 and S_val > 0:
            if exercise < S_val * 0.01:   # tỷ lệ < 0.01 → khả năng exercise ở dạng nghìn đồng
                exercise_vnd = exercise * 1000

        if S_val > 0 and exercise_vnd > 0 and T > 0 and sigma > 0:
            bs_call = black_scholes_option(S_val, exercise_vnd, T, risk_free, sigma, ratio, 'call')
            bs_put  = black_scholes_option(S_val, exercise_vnd, T, risk_free, sigma, ratio, 'put')
        else:
            pass

        # BUG FIX #3: Thêm field 'delta' chuẩn để frontend dùng được (chọn theo option_type)
        delta_effective = bs_call["delta_bs"] if is_call else abs(bs_put["delta_bs"])
        # Effective Gearing = (S / (price_cw_vnd * ratio)) * delta_option  [đã kiểm tra đúng]
        eff_lev_val = round(gross_lev * delta_effective, 2) if gross_lev > 0 else 0.0

        cw_list.append({
            "ticker": ticker,
            "underlying": und_sym,
            "issuer": str(row.get("to_chuc_ph_cw","")).replace("CTCP Chứng khoán ","").split("(")[0].strip(),
            "maturity": str(row.get("ngay_dao_han","")),
            "status": "active" if pd.notna(row["_last_gd_dt"]) and row["_last_gd_dt"] >= today_ts else "expired",
            "option_type": option_type,
            "underlying_price": S_val,
            "price": round(price_cw, 3),
            "exercise": exercise_vnd,   # FIX #5: luôn là VND gốc
            "ratio": ratio,
            "premium": round(premium, 1),
            "eff_lev": eff_lev_val,
            "breakeven": round(breakeven, 1),
            "moneyness": moneyness,
            # BUG FIX #3: Thêm 'delta' field chuẩn (|delta| của option_type thực tế) để frontend dùng
            "delta": round(delta_effective, 4),
            # Kết quả Black‑Scholes đầy đủ
            "bs_price_call": bs_call["bs_price"],
            "delta_call": bs_call["delta_bs"],
            "bs_price_put": bs_put["bs_price"],
            "delta_put": bs_put["delta_bs"],
            "gamma": bs_call["gamma"],    # gamma giống nhau call/put
            "vega": bs_call["vega"],
            "theta": bs_call["theta"],    # theta của call (để tham khảo)
            "sigma": round(sigma, 4),
            "risk_free": round(risk_free, 4),
            # Thêm d1, d2 để dashboard có thể vẽ BS visualization
            "d1": round((log(S_val / exercise_vnd) + (risk_free + 0.5 * sigma**2) * T) / (sigma * sqrt(T)), 4) if (S_val > 0 and exercise_vnd > 0 and T > 0 and sigma > 0) else 0,
            "T_years": round(T, 4),
        })

    # --- OHLCV drill-down cho dashboard (bao gồm giá cơ sở) ---
    ohlcv_out = {}
    for ticker in valid_tickers:
        sub = ohlcv_idx.get(ticker, pd.DataFrame())
        if sub.empty:
            continue
        sub = sub.sort_index()
        und_sym = df_info.loc[df_info["ma_cw"] == ticker, "ck_co_so"]
        und_sym = und_sym.values[0] if len(und_sym) else ""

        # BUG FIX EG: Build underlying_close bằng reindex + ffill thay vì dict lookup
        # Vấn đề cũ: und_ser.get(d, 0) → trả 0 khi ngày không khớp chính xác
        # (CW giao dịch vào ngày mà underlying không có dữ liệu, hoặc format lệch)
        df_und = underlying_ohlcv.get(und_sym, pd.DataFrame())
        if not df_und.empty and "close" in df_und.columns:
            # Reindex underlying theo index của CW, forward-fill để điền ngày thiếu
            und_aligned = df_und["close"].reindex(sub.index, method="ffill")
            underlying_close = []
            for v in und_aligned:
                try:
                    price = float(v)
                    if price > 0 and price < 1000:
                        price *= 1000
                    underlying_close.append(round(price) if price > 0 else 0)
                except Exception:
                    underlying_close.append(0)
        else:
            # Fallback: thử dùng dict (backward compat)
            und_ser = underlying_price_series.get(und_sym, {})
            dates_str = sub.index.strftime("%d/%m/%Y").tolist()
            underlying_close = [und_ser.get(d, 0) for d in dates_str]

        dates = sub.index.strftime("%d/%m/%Y").tolist()
        entry = {
            "dates": dates,
            "close": [round(float(v), 3) for v in sub["close"]],
            "underlying_close": underlying_close,
        }
        for col in ["open", "high", "low"]:
            if col in sub.columns and sub[col].notna().any():
                entry[col] = [round(float(v), 3) if pd.notna(v) else None for v in sub[col]]
        if "volume" in sub.columns and sub["volume"].notna().any():
            entry["volume"] = [int(v) if pd.notna(v) else 0 for v in sub["volume"]]
        ohlcv_out[ticker] = entry

    out = {
        # BUG FIX #4: Lưu ICT trực tiếp với suffix rõ ràng "+07:00" để frontend parse đúng, không cộng thêm 7h
        "updated_at": (datetime.now(timezone.utc) + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M +07:00"),
        "cw_list": cw_list,
        "ohlcv": ohlcv_out,
    }
    os.makedirs("docs", exist_ok=True)
    with open("docs/data.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize("docs/data.json") / 1024
    print(f"OK docs/data.json  ({len(cw_list)} CW, {len(ohlcv_out)} OHLCV, {size_kb:.0f} KB)")
    print("\n   Sample metrics (3 CW đầu):")
    for c in cw_list[:3]:
        print(f"   {c['ticker']:12} S={c.get('underlying_price',0):>8,.0f}  "
              f"K={c['exercise']:>8,}  Premium={c['premium']:>6.1f}%  "
              f"BS_call={c['bs_price_call']:.3f}  BS_put={c['bs_price_put']:.3f}  "
              f"{c['moneyness']}  σ={c['sigma']:.3f}")

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    t0 = time.time()
    df_vs              = step1_vietstock()
    df_ohlcv_full      = step2_ohlcv(df_vs)
    df_filtered, valid = step3_filter(df_ohlcv_full)
    step4_excel(df_filtered, df_vs, valid)
    step5_export_json(df_ohlcv_full, df_vs, valid)
    print(f"\nDone in {(time.time()-t0)/60:.1f} min")
